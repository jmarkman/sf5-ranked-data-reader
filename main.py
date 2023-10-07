import json
import openpyxl
import dateutil.parser
from datatypes import RankedMatch, RankedSession, GameResult
from encoder import RankedSessionEncoder

def read_data_from_workbook():
    ranked_data_workbook = openpyxl.load_workbook("jivegrind.xlsx")
    sheets = [x for x in ranked_data_workbook.sheetnames if "New" not in x]
    sessions = []
    for sheet in sheets:
        worksheet = ranked_data_workbook[sheet]
        matches = []
        for row in worksheet.iter_rows(min_row=2, max_col=4, values_only=True):
            if not all(row):
                print(f"Encountered empty row in sheet '{sheet}', reached end of matches")
                break
            else:
                matches.append(create_ranked_match(row))
        sessions.append(create_ranked_session(worksheet, matches))
    return sessions

def create_ranked_match(row_data):
    if all(row_data):
        points = int(row_data[0])
        result = GameResult.from_str(row_data[1])
        opponent = row_data[2]
        replay_id = row_data[3]
        return RankedMatch(points, result, opponent, replay_id)

def create_ranked_session(worksheet, matches):
    session_date = str(dateutil.parser.parse(worksheet.title).date())
    points_start = int(worksheet["F2"].value)
    points_end = int(worksheet["G2"].value)
    return RankedSession(session_date, points_start, points_end, matches)

if __name__ == '__main__':
    all_ranked_sessions = read_data_from_workbook()
    ranked_sessions_json = json.dumps(all_ranked_sessions, cls=RankedSessionEncoder)
    with open("ranked-data.json", 'w', encoding='utf-8') as writer:
        print("Dumping serialized ranked sessions")
        json.dump(all_ranked_sessions, writer, ensure_ascii=False, cls=RankedSessionEncoder)